"""
JANSA VISASIST — Excel → list[WorkflowRow] loader.
Includes SOCOTEC verdict injection for Bureau de Contrôle "En attente" rows.
"""
import json
from typing import Optional

import openpyxl

from processing.config import (
    TARGET_SHEET, DATA_START_ROW,
    COL_MAP_METADATA, COL_MAP_WORKFLOW,
    make_submittal_key, TODAY,
    is_actor_moex,
    DISCARD_EMPTY_ACTOR, ANCIEN_EXCEPTION_SUBMITTAL_IDS,
    GOE_COMPLETED_EXCEPTION_IDS,
    SOCOTEC_VERDICTS_PATH,
)
from processing.models import WorkflowRow, Anomaly
from processing.dates import parse_date, parse_delay
from processing.actors import resolve_actor
from processing.statuses import resolve_status


def _load_socotec_verdicts() -> dict[str, dict]:
    """Load hardcoded SOCOTEC verdicts from data/socotec_verdicts.json.
    Returns {numero: {avis_raw, avis_excel, avis_code, observation, fiche, fiche_date}}.
    """
    if not SOCOTEC_VERDICTS_PATH.exists():
        return {}
    with open(SOCOTEC_VERDICTS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def _col_idx(letter: str) -> int:
    """Convert column letter (A, Z, AA, AF…) to 1-based column index."""
    r = 0
    for c in letter:
        r = r * 26 + (ord(c.upper()) - ord('A') + 1)
    return r


def load_workbook(
    excel_path: str,
    actor_map: dict,
    status_map: dict,
) -> tuple[list[WorkflowRow], list[Anomaly]]:
    """Load Excel file and build WorkflowRow list.

    Returns (rows, anomalies).
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[TARGET_SHEET]

    meta_cols = {_col_idx(l): f for l, f in COL_MAP_METADATA.items()}
    wf_cols = {_col_idx(l): f for l, f in COL_MAP_WORKFLOW.items()}

    # Load SOCOTEC hardcoded verdicts
    socotec_verdicts = _load_socotec_verdicts()
    socotec_injected = 0

    rows: list[WorkflowRow] = []
    anomalies: list[Anomaly] = []
    unmapped_actors: set[str] = set()
    unknown_tags: set[str] = set()
    discarded_empty_actor = 0
    discarded_ancien = 0
    discarded_goe_completed = 0

    for row_i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW), start=DATA_START_ROW):
        cells: dict[int, object] = {}
        for c in row:
            try:
                cells[c.column] = c.value
            except AttributeError:
                pass

        meta = {f: cells.get(ci) for ci, f in meta_cols.items()}
        wf = {f: cells.get(ci) for ci, f in wf_cols.items()}

        # --- Exception Type 1: Discard rows with empty actor (col Z) ---
        actor_raw = wf.get("actor_raw")
        if DISCARD_EMPTY_ACTOR and (actor_raw is None or str(actor_raw).strip() == ""):
            discarded_empty_actor += 1
            continue

        # --- Exception Type 2: Discard ancien submittal IDs ---
        sid_raw = meta.get("submittal_id")
        sid_str = str(sid_raw) if sid_raw is not None else None
        if sid_str and sid_str in ANCIEN_EXCEPTION_SUBMITTAL_IDS:
            discarded_ancien += 1
            continue

        # --- Exception Type 3: Discard GOE completed submittals ---
        if sid_str and sid_str in GOE_COMPLETED_EXCEPTION_IDS:
            discarded_goe_completed += 1
            continue

        # --- Actor resolution ---
        ae, actor_mapped = resolve_actor(actor_raw, actor_map)
        if not actor_mapped:
            akey = str(actor_raw).strip() if actor_raw is not None else "_UNKNOWN_"
            if akey not in unmapped_actors:
                unmapped_actors.add(akey)
            anomalies.append(Anomaly(
                row_i, "unmapped_actor", "actor_raw", akey,
                f"Actor '{akey}' not in map → irrelevant",
            ))

        # --- Status resolution ---
        tag_raw = wf.get("response_tag_raw")
        se, tag_mapped = resolve_status(tag_raw, status_map)
        if not tag_mapped:
            tkey = str(tag_raw).strip() if tag_raw is not None else "_NONE_"
            if tkey not in unknown_tags:
                unknown_tags.add(tkey)
            anomalies.append(Anomaly(
                row_i, "unknown_status_tag", "response_tag_raw", tkey,
                f"Unknown tag '{tkey}'",
            ))

        # --- Parse fields ---
        dl = parse_date(wf.get("deadline_raw"))
        rd = parse_date(wf.get("response_date_raw"))
        tc, sv = se["code"], se["severity"]
        rel = ae["relevant"]
        moex_flag = is_actor_moex(ae)

        # --- SOCOTEC verdict injection ---
        # If this is a Bureau de Contrôle row with "En attente" status and
        # the numero matches a hardcoded SOCOTEC verdict, inject it.
        socotec_applied = False
        numero_str = str(meta["numero"]).strip() if meta.get("numero") is not None else ""
        actor_str = str(actor_raw).strip() if actor_raw else ""

        if (socotec_verdicts
                and "Bureau de Contrôle" in actor_str
                and tc == "EN_ATTENTE"
                and numero_str in socotec_verdicts):
            sv_entry = socotec_verdicts[numero_str]
            # Override status fields
            tag_raw = sv_entry["avis_excel"]
            tc = sv_entry["avis_code"]
            # Map avis_code to severity
            _sev_map = {"FAV": "favorable", "SUSP": "caution", "DEF": "blocking"}
            sv = _sev_map.get(tc, "non_response")
            # Override response date with fiche date
            if sv_entry.get("fiche_date"):
                rd = parse_date(sv_entry["fiche_date"])
                wf["response_date_raw"] = sv_entry["fiche_date"]
            # Override respondant and comment
            wf["respondant"] = "SOCOTEC (PDF)"
            if sv_entry.get("observation"):
                wf["comment_raw"] = sv_entry["observation"]
            # Re-resolve the status entry for downstream use
            se = {"code": tc, "severity": sv}
            socotec_applied = True
            socotec_injected += 1

        # --- Build WorkflowRow ---
        wr = WorkflowRow(
            source_row_index=row_i,
            directory=meta.get("directory"),
            submittal_id=str(meta["submittal_id"]) if meta.get("submittal_id") is not None else None,
            affaire=meta.get("affaire"),
            projet=meta.get("projet"),
            batiment=meta.get("batiment"),
            phase=meta.get("phase"),
            emetteur=meta.get("emetteur"),
            specialite=meta.get("specialite"),
            lot=meta.get("lot"),
            type_doc=meta.get("type_doc"),
            zone=meta.get("zone"),
            niveau=meta.get("niveau"),
            numero=str(meta["numero"]) if meta.get("numero") is not None else None,
            indice=meta.get("indice"),
            attached_filename=meta.get("attached_filename"),
            submittal_key=make_submittal_key(meta.get("submittal_id"), meta.get("indice")),
            actor_raw=actor_raw,
            actor_prefix=ae["prefix"],
            actor_role=ae["role"],
            actor_clean=ae["canonical"],
            actor_family=ae["family"],
            is_relevant_actor=rel,
            is_moex=moex_flag,
            respondant=wf.get("respondant"),
            deadline_raw=wf.get("deadline_raw"),
            deadline_date=dl,
            response_date_raw=wf.get("response_date_raw"),
            response_date=rd,
            delay_raw=wf.get("delay_raw"),
            delay_days=parse_delay(wf.get("delay_raw")),
            response_tag_raw=tag_raw,
            response_tag_clean=str(tag_raw).strip() if tag_raw is not None else "_NONE_",
            response_tag_code=tc,
            response_severity=sv,
            comment_raw=wf.get("comment_raw"),
            is_pending=(tc == "EN_ATTENTE"),
            is_late=(tc == "EN_ATTENTE" and dl is not None and dl < TODAY),
            is_hors_mission=(tc == "HM"),
            has_response_raw=(tc != "NONE"),
            has_effective_response=(tc not in ("EN_ATTENTE", "NONE") and sv not in ("neutral", "non_response")),
            is_active_row=(rel and tc not in ("HM", "NONE") and sv != "neutral"),
        )
        rows.append(wr)

    wb.close()

    # Attach SOCOTEC injection stats
    if socotec_injected > 0:
        anomalies.append(Anomaly(
            0, "socotec_verdict_injection", "response_tag_raw",
            str(socotec_injected),
            f"Injected SOCOTEC verdicts into {socotec_injected} BdC 'En attente' rows "
            f"(from {len(socotec_verdicts)} hardcoded verdicts)",
        ))

    # Attach discard stats as anomalies for reporting
    if discarded_empty_actor > 0:
        anomalies.append(Anomaly(
            0, "exception_type1_empty_actor", "actor_raw", str(discarded_empty_actor),
            f"Discarded {discarded_empty_actor} rows with empty actor (col Z)",
        ))
    if discarded_ancien > 0:
        anomalies.append(Anomaly(
            0, "exception_type2_ancien_submittal", "submittal_id", str(discarded_ancien),
            f"Discarded {discarded_ancien} rows belonging to {len(ANCIEN_EXCEPTION_SUBMITTAL_IDS)} ancien submittal IDs",
        ))
    if discarded_goe_completed > 0:
        anomalies.append(Anomaly(
            0, "exception_type3_goe_completed", "submittal_id", str(discarded_goe_completed),
            f"Discarded {discarded_goe_completed} rows belonging to {len(GOE_COMPLETED_EXCEPTION_IDS)} GOE completed submittal IDs",
        ))

    return rows, anomalies
