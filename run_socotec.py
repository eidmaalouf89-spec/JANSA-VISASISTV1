"""
JANSA VISASIST — SOCOTEC PDF → GED Excel Pipeline
Usage: python run_socotec.py

Reads:  socotec reports/*.pdf
Updates: source/17&CO Tranche 2 du 23 mars 2026 07_45.xlsx
Output: source/17&CO Tranche 2 du 23 mars 2026 07_45_UPDATED.xlsx  (new copy)

Persists:
  data/socotec_verdicts.json  — hardcoded verdicts (merged with existing)
  data/socotec_registry.json  — PDF processing registry (appended)

Deterministic rules:
  1. Parse all SOCOTEC "Fiche examen de document" PDFs
  2. Extract document codes (P17_T2_...) + verdicts (F/S/D) + observations
  3. Match by 6-digit numero to column M of the GED Excel
  4. Only update Bureau de Contrôle rows (col Z) where tag = "En attente" (col AE)
  5. Later fiche overwrites earlier fiche for same numero
  6. Never touch rows that already have a response
  7. Persist verdicts → data/socotec_verdicts.json (new PDFs merge into existing)
  8. Persist registry → data/socotec_registry.json (track processed PDFs)
"""
import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber

# ═══════════════════════════════════════════════════════════════════════════
# Config
# ═══════════════════════════════════════════════════════════════════════════

PROJECT_ROOT = Path(__file__).resolve().parent
PDF_DIR = PROJECT_ROOT / "socotec reports"
SOURCE_DIR = PROJECT_ROOT / "source"
EXCEL_NAME = "17&CO Tranche 2 du 23 mars 2026 07_45.xlsx"
EXCEL_PATH = SOURCE_DIR / EXCEL_NAME
OUTPUT_PATH = SOURCE_DIR / EXCEL_NAME.replace(".xlsx", "_UPDATED.xlsx")

SHEET_NAME = "Vue détaillée des documents 1"
DATA_START_ROW = 3

# Column positions (1-based)
COL_M_NUMERO = 13
COL_Z_ACTOR = 26
COL_AA_RESPONDANT = 27
COL_AC_RESPONSE_DATE = 29
COL_AE_RESPONSE_TAG = 31
COL_AF_COMMENT = 32

# Verdict mapping: PDF avis → Excel tag value
AVIS_TO_EXCEL = {
    "F": "Favorable",
    "S": "Suspendu",
    "D": "Défavorable",
}

# Document code regex
DOC_CODE_RE = re.compile(
    r'P17_T2_(\w+)_EXE_(\w+)_(\w+)_(\w+)_(\w+)_(\w+)_(\w+)_(\d{6})_(\w+)-(\w+)'
)
FICHE_NUM_RE = re.compile(r'Fiche\s+n[°o]?\s*[:.]?\s*(\d+)', re.IGNORECASE)
FILENAME_DATE_RE = re.compile(r'(\d{2})-(\d{2})-(\d{2})')

# Data persistence paths
DATA_DIR = PROJECT_ROOT / "data"
VERDICTS_JSON = DATA_DIR / "socotec_verdicts.json"
REGISTRY_JSON = DATA_DIR / "socotec_registry.json"

# Avis code mapping (for verdicts JSON persistence)
AVIS_TO_CODE = {"F": "FAV", "S": "SUSP", "D": "DEF"}


# ═══════════════════════════════════════════════════════════════════════════
# PDF text normalization
# ═══════════════════════════════════════════════════════════════════════════

def _norm(text: str) -> str:
    """Collapse newlines that break P17 document codes in PDF tables."""
    text = re.sub(r'_\s*\n(\w)', r'_\1', text)
    text = re.sub(r'_(\w)\s*\n-(\w)', r'_\1-\2', text)
    return text


def _parse_filename_date(filename: str) -> str | None:
    """Extract DD/MM/YYYY from filename format DD-MM-YY."""
    m = FILENAME_DATE_RE.search(filename)
    if m:
        dd, mm, yy = m.groups()
        return f"{dd}/{mm}/20{yy}"
    return None


# ═══════════════════════════════════════════════════════════════════════════
# Step 1: Parse PDFs
# ═══════════════════════════════════════════════════════════════════════════

def parse_pdfs(pdf_dir: Path) -> tuple[dict, list[dict]]:
    """Parse all SOCOTEC PDFs.

    Returns:
        verdicts: {numero → {avis, avis_excel, observation, fiche, fiche_date, full_code}}
        reports:  [{ filename, fiche, fiche_date, pages, docs_listed, verdicts_explicit, ... }]
    """
    pdf_files = sorted(pdf_dir.glob("*.pdf"))
    if not pdf_files:
        print(f"  [ERROR] No PDFs found in {pdf_dir}")
        return {}, []

    # Per-report tracking
    reports = []
    # All verdicts, keyed by numero. Later fiche overwrites earlier.
    all_verdicts: dict[str, dict] = {}
    # All listed doc numeros (for implicit FAV)
    all_listed: dict[str, dict] = {}

    for pdf_path in pdf_files:
        fiche_num = None
        fiche_date = _parse_filename_date(pdf_path.name)
        listed_numeros = set()
        explicit_verdicts = {}
        report_docs = 0

        with pdfplumber.open(pdf_path) as pdf:
            # Fiche number from first page
            p1 = pdf.pages[0].extract_text() or ""
            fm = FICHE_NUM_RE.search(p1)
            if fm:
                fiche_num = fm.group(1)

            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(c).strip() if c else "" for c in table[0]]

                    # ── Document list table ──
                    if "Désignation" in header[0]:
                        for row in table[1:]:
                            cell = _norm(str(row[0])) if row[0] else ""
                            for m in DOC_CODE_RE.finditer(cell):
                                num = m.group(8)
                                listed_numeros.add(num)
                                report_docs += 1
                                if num not in all_listed:
                                    all_listed[num] = {
                                        "full_code": m.group(0),
                                        "fiche": fiche_num,
                                        "fiche_date": fiche_date,
                                    }

                    # ── Verdict table ──
                    elif ("léments examinés" in header[0]
                          or (len(header) > 1 and "Avis" in header[1])):
                        for row in table[1:]:
                            elem = _norm(str(row[0])) if row[0] else ""
                            avis = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                            obs = str(row[2]).strip() if len(row) > 2 and row[2] else ""

                            if avis not in ("F", "S", "D"):
                                continue

                            codes = DOC_CODE_RE.findall(elem)
                            for c in codes:
                                num = c[7]
                                full = f"P17_T2_{c[0]}_EXE_{c[1]}_{c[2]}_{c[3]}_{c[4]}_{c[5]}_{c[6]}_{c[7]}_{c[8]}-{c[9]}"
                                explicit_verdicts[num] = {
                                    "avis": avis,
                                    "avis_excel": AVIS_TO_EXCEL[avis],
                                    "observation": obs,
                                    "fiche": fiche_num,
                                    "fiche_date": fiche_date,
                                    "full_code": full,
                                }
                                listed_numeros.add(num)

        # Merge into global verdicts (later fiche overwrites)
        # Explicit verdicts first
        for num, v in explicit_verdicts.items():
            all_verdicts[num] = v

        # Docs listed but NOT in explicit verdicts → implicit Favorable
        for num in listed_numeros:
            if num not in all_verdicts:
                all_verdicts[num] = {
                    "avis": "F",
                    "avis_excel": "Favorable",
                    "observation": "(implicite — aucune observation BdC)",
                    "fiche": fiche_num,
                    "fiche_date": fiche_date,
                    "full_code": all_listed.get(num, {}).get("full_code", ""),
                }

        reports.append({
            "filename": pdf_path.name,
            "fiche": fiche_num,
            "fiche_date": fiche_date,
            "pages": len(pdfplumber.open(pdf_path).pages),
            "docs_listed": report_docs,
            "verdicts_explicit": len(explicit_verdicts),
            "verdicts_f": sum(1 for v in explicit_verdicts.values() if v["avis"] == "F"),
            "verdicts_s": sum(1 for v in explicit_verdicts.values() if v["avis"] == "S"),
            "verdicts_d": sum(1 for v in explicit_verdicts.values() if v["avis"] == "D"),
        })

    return all_verdicts, reports


# ═══════════════════════════════════════════════════════════════════════════
# Step 2: Update Excel
# ═══════════════════════════════════════════════════════════════════════════

def update_excel(
    excel_in: Path,
    excel_out: Path,
    verdicts: dict[str, dict],
) -> dict:
    """Open the GED Excel, update BdC 'En attente' rows, save a new copy.

    Returns stats dict.
    """
    # Copy first — never modify original
    shutil.copy2(excel_in, excel_out)

    wb = openpyxl.load_workbook(excel_out)
    ws = wb[SHEET_NAME]

    stats = {
        "bdc_total": 0,
        "bdc_en_attente": 0,
        "updated": 0,
        "skipped_already_responded": 0,
        "skipped_no_pdf_match": 0,
        "updated_rows": [],
    }

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
        # Build cell dict
        cells = {}
        for c in row:
            cells[c.column] = c

        actor_cell = cells.get(COL_Z_ACTOR)
        actor = str(actor_cell.value).strip() if actor_cell and actor_cell.value else ""

        if "Bureau de Contrôle" not in actor:
            continue

        stats["bdc_total"] += 1

        tag_cell = cells.get(COL_AE_RESPONSE_TAG)
        current_tag = str(tag_cell.value).strip() if tag_cell and tag_cell.value else ""

        if current_tag != "En attente":
            stats["skipped_already_responded"] += 1
            continue

        stats["bdc_en_attente"] += 1

        num_cell = cells.get(COL_M_NUMERO)
        numero = str(num_cell.value).strip() if num_cell and num_cell.value else ""

        if numero not in verdicts:
            stats["skipped_no_pdf_match"] += 1
            continue

        # ── UPDATE ──
        v = verdicts[numero]

        # Col AE: response tag
        tag_cell.value = v["avis_excel"]

        # Col AC: response date (use fiche date)
        date_cell = cells.get(COL_AC_RESPONSE_DATE)
        if date_cell and v["fiche_date"]:
            date_cell.value = v["fiche_date"]

        # Col AA: respondant
        resp_cell = cells.get(COL_AA_RESPONDANT)
        if resp_cell:
            resp_cell.value = "SOCOTEC (PDF)"

        # Col AF: comment/observation
        comment_cell = cells.get(COL_AF_COMMENT)
        if comment_cell and v["observation"]:
            comment_cell.value = v["observation"]

        stats["updated"] += 1
        stats["updated_rows"].append({
            "row": row[0].row,
            "numero": numero,
            "verdict": v["avis_excel"],
            "fiche": v["fiche"],
            "observation": v["observation"][:80] if v["observation"] else "",
        })

    wb.save(excel_out)
    wb.close()
    return stats


# ═══════════════════════════════════════════════════════════════════════════
# Step 3: Persist verdicts & registry to data/
# ═══════════════════════════════════════════════════════════════════════════

def persist_verdicts(verdicts: dict[str, dict]) -> dict:
    """Merge new verdicts into data/socotec_verdicts.json.

    Returns stats: {existing, new_added, updated, total_after}.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    existing: dict[str, dict] = {}
    if VERDICTS_JSON.exists():
        with open(VERDICTS_JSON, "r", encoding="utf-8") as f:
            existing = json.load(f)

    stats = {"existing": len(existing), "new_added": 0, "updated": 0}

    for num, v in verdicts.items():
        entry = {
            "avis_raw": v["avis"],
            "avis_excel": v["avis_excel"],
            "avis_code": AVIS_TO_CODE.get(v["avis"], v["avis"]),
            "observation": v.get("observation", ""),
            "fiche": v.get("fiche", ""),
            "fiche_date": v.get("fiche_date", ""),
        }
        if num in existing:
            stats["updated"] += 1
        else:
            stats["new_added"] += 1
        existing[num] = entry

    stats["total_after"] = len(existing)

    with open(VERDICTS_JSON, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    return stats


def persist_registry(reports: list[dict]) -> dict:
    """Merge new report entries into data/socotec_registry.json.

    Deduplicates by fiche_number, keeping the entry with a date.
    Returns stats: {existing, new_added, total_after}.
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    existing: list[dict] = []
    if REGISTRY_JSON.exists():
        with open(REGISTRY_JSON, "r", encoding="utf-8") as f:
            existing = json.load(f)

    # Index existing by fiche_number
    by_fiche: dict[str, dict] = {}
    for entry in existing:
        fn = entry.get("fiche_number", "")
        if fn:
            by_fiche[fn] = entry

    stats = {"existing": len(by_fiche), "new_added": 0}

    for r in reports:
        fn = r.get("fiche", "")
        if not fn:
            continue
        entry = {
            "filename": r["filename"],
            "fiche_number": fn,
            "fiche_date": r.get("fiche_date", ""),
            "pages": r.get("pages", 0),
            "docs_listed": r.get("docs_listed", 0),
            "explicit_verdicts": r.get("verdicts_explicit", 0),
            "fav_explicit": r.get("verdicts_f", 0),
            "susp_explicit": r.get("verdicts_s", 0),
            "def_explicit": r.get("verdicts_d", 0),
        }
        if fn not in by_fiche:
            stats["new_added"] += 1
        # Prefer entry with date
        if fn in by_fiche and not by_fiche[fn].get("fiche_date") and entry["fiche_date"]:
            by_fiche[fn] = entry
        elif fn not in by_fiche:
            by_fiche[fn] = entry

    # Sort by fiche_number
    result = sorted(by_fiche.values(), key=lambda x: int(x.get("fiche_number", "0") or "0"))
    stats["total_after"] = len(result)

    with open(REGISTRY_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return stats


# ═══════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════

def main():
    P = print

    P("=" * 74)
    P("  JANSA VISASIST — SOCOTEC PDF PIPELINE")
    P("=" * 74)

    # ── Step 1: Parse PDFs ──
    P(f"\n{'─'*74}\n  1. PDF EXTRACTION\n{'─'*74}")
    P(f"  Source: {PDF_DIR}")

    if not PDF_DIR.exists():
        P(f"  [ERROR] Folder not found: {PDF_DIR}")
        sys.exit(1)

    verdicts, reports = parse_pdfs(PDF_DIR)

    P(f"  Reports found: {len(reports)}")
    for r in reports:
        P(f"    Fiche #{r['fiche']} ({r['fiche_date']}): {r['pages']}p, "
          f"{r['docs_listed']} docs, "
          f"F={r['verdicts_f']} S={r['verdicts_s']} D={r['verdicts_d']} explicit")

    P(f"\n  Total unique verdicts: {len(verdicts)}")
    from collections import Counter
    vc = Counter(v["avis_excel"] for v in verdicts.values())
    for tag, cnt in vc.most_common():
        P(f"    {tag}: {cnt}")

    # ── Step 2: Update Excel ──
    P(f"\n{'─'*74}\n  2. EXCEL UPDATE\n{'─'*74}")
    P(f"  Input:  {EXCEL_PATH}")
    P(f"  Output: {OUTPUT_PATH}")

    if not EXCEL_PATH.exists():
        P(f"  [ERROR] Excel not found: {EXCEL_PATH}")
        sys.exit(1)

    stats = update_excel(EXCEL_PATH, OUTPUT_PATH, verdicts)

    P(f"\n  Bureau de Contrôle rows total: {stats['bdc_total']}")
    P(f"  Already responded (skipped):   {stats['skipped_already_responded']}")
    P(f"  En attente:                    {stats['bdc_en_attente']}")
    P(f"  No PDF match (skipped):        {stats['skipped_no_pdf_match']}")
    P(f"  ✓ UPDATED:                     {stats['updated']}")

    # ── Step 3: Persist to data/ ──
    P(f"\n{'─'*74}\n  3. PERSIST TO data/\n{'─'*74}")
    v_stats = persist_verdicts(verdicts)
    P(f"  Verdicts JSON: {VERDICTS_JSON}")
    P(f"    Existing: {v_stats['existing']}  New: {v_stats['new_added']}  Updated: {v_stats['updated']}  Total: {v_stats['total_after']}")

    r_stats = persist_registry(reports)
    P(f"  Registry JSON: {REGISTRY_JSON}")
    P(f"    Existing: {r_stats['existing']}  New: {r_stats['new_added']}  Total: {r_stats['total_after']}")

    # ── Step 4: Detail of updates ──
    P(f"\n{'─'*74}\n  4. UPDATED ROWS DETAIL\n{'─'*74}")
    if stats["updated_rows"]:
        P(f"  {'Row':>6s}  {'Numéro':>8s}  {'Verdict':<14s}  {'Fiche':>5s}  Observation")
        P(f"  {'─'*6}  {'─'*8}  {'─'*14}  {'─'*5}  {'─'*40}")
        for u in stats["updated_rows"]:
            P(f"  {u['row']:>6d}  {u['numero']:>8s}  {u['verdict']:<14s}  "
              f"#{u['fiche']:>4s}  {u['observation']}")
    else:
        P("  No rows updated.")

    # ── Step 5: Summary ──
    P(f"\n{'─'*74}\n  5. SUMMARY\n{'─'*74}")
    remaining = stats["bdc_en_attente"] - stats["updated"]
    P(f"  BdC 'En attente' before: {stats['bdc_en_attente']}")
    P(f"  BdC 'En attente' after:  {remaining}")
    P(f"  Rows updated:            {stats['updated']}")
    P(f"  Remaining unresolved:    {remaining} ({remaining} have no SOCOTEC PDF coverage yet)")
    P(f"\n  Updated file saved to:")
    P(f"  → {OUTPUT_PATH}")

    P(f"\n{'═'*74}")
    P(f"  SOCOTEC PIPELINE COMPLETE")
    P(f"{'═'*74}")


if __name__ == "__main__":
    main()
